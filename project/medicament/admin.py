import openpyxl
from django.http import HttpResponse
from django.contrib import admin
from .models import Medicament
from datetime import date, timedelta, datetime
from django.db.models import Sum

from django.contrib import admin
from .models import Medicament

from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import redirect
from .models import Medicament,Medicament2
from .oracle2 import update_medicaments
from .oracle2copy import update_medicaments2

@admin.register(Medicament2)
class Medicament2Admin(admin.ModelAdmin):
    search_fields = ('name',)
    actions = ['export_to_excel', 'export_all_to_excel']
    list_display = (
    "med_id",
    "name",
    "qte",
    "valeur_achat",)

    change_list_template = "admin/Medicament_perimes_personnalise.html"

    def sync_oracle2(self, request):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        start_date = start_date.strftime("%d/%m/%Y")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date.strftime("%d/%m/%Y")
        try:
            update_medicaments2(start_date, end_date)
            self.message_user(
                request,
                "Oracle data synchronized successfully!",
                level=messages.SUCCESS,
            )
        except Exception as e:
            self.message_user(request, f"Error: {str(e)}", level=messages.ERROR)
        return redirect("../")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "sync-oracle2/",
                self.admin_site.admin_view(self.sync_oracle2),
                name="sync_oracle2",
            ),
        ]
        return custom_urls + urls

@admin.register(Medicament)
class MedicamentAdmin(admin.ModelAdmin):
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('sync-oracle/', self.admin_site.admin_view(self.sync_oracle), name="sync_oracle"),
            path("export-excel/", self.admin_site.admin_view(self._export_to_excel), name="medicament_export_to_excel"),
        ]
        return custom_urls + urls
    
    search_fields = ('name',)
    actions = ['export_to_excel', 'export_all_to_excel']
    list_display = (
    "med_id",
    "name",
    "qte_total",
    "val_total_achat",

    # < 3 mois
    "qte_avant_3m",
    "val_achat_avant_3m",

    # 3–6 mois
    "qte_apres_3m_avant_6m",
    "val_achat_apres_3m_avant_6m",

    # 6–12 mois
    "qte_apres_6m_avant_12m",
    "val_achat_apres_6m_avant_12m",

    # 12–18 mois
    "qte_apres_12m_avant_18m",
    "val_achat_apres_12m_avant_18m",

    # 18–24 mois
    "qte_apres_18m_avant_24m",
    "val_achat_apres_18m_avant_24m",

    # 24–36 mois
    "qte_apres_24m_avant_36m",
    "val_achat_apres_24m_avant_36m",

    # > 36 mois
    "qte_apres_36m",
    "val_achat_apres_36m",
    )

    change_list_template = "admin/medicament_changelist.html"

    class Media:
        js = ("admin/js/ttt.js",)


    def sync_oracle(self, request):
        try:
            update_medicaments()
            self.message_user(request, "Oracle data synchronized successfully!", level=messages.SUCCESS)
        except Exception as e:
            self.message_user(request, f"Error: {str(e)}", level=messages.ERROR)
        return redirect("../")
    
    # -------------------- Export Action --------------------
    def export_all_to_excel(self, request, queryset):
        all_medicaments = Medicament.objects.all()
        return self._export_to_excel(all_medicaments)

    export_all_to_excel.short_description = "Exporter tous les médicaments vers Excel"
    
    def _export_to_excel(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Medicaments"

        queryset = Medicament.objects.all()

        # Data rows
        for obj in queryset:
            row = [
                    obj.med_id,
                    obj.name,

                    # Totaux
                    obj.qte_total,
                    obj.val_total_achat,

                    # < 3 mois
                    obj.qte_avant_3m,
                    obj.val_achat_avant_3m,

                    # 3–6 mois
                    obj.qte_apres_3m_avant_6m,
                    obj.val_achat_apres_3m_avant_6m,

                    # 6–12 mois
                    obj.qte_apres_6m_avant_12m,
                    obj.val_achat_apres_6m_avant_12m,

                    # 12–18 mois
                    obj.qte_apres_12m_avant_18m,
                    obj.val_achat_apres_12m_avant_18m,

                    # 18–24 mois
                    obj.qte_apres_18m_avant_24m,
                    obj.val_achat_apres_18m_avant_24m,

                    # 24–36 mois
                    obj.qte_apres_24m_avant_36m,
                    obj.val_achat_apres_24m_avant_36m,

                    # > 36 mois
                    obj.qte_apres_36m,
                    obj.val_achat_apres_36m,
                ]
            sheet.append(row)
        
        sheet.insert_rows(1)
        sheet.insert_rows(1)
        sheet.merge_cells("A1:A2")
        sheet["A1"] = "Med ID"
        sheet.merge_cells("B1:B2")
        sheet["B1"] = "Medicament"
        sheet.merge_cells("C1:D1")
        sheet["C1"] = "TOTAL"
        sheet.merge_cells("E1:F1")
        sheet["E1"] = "< 3 mois"
        
        sheet.merge_cells("G1:H1")
        sheet["G1"] = "3–6 mois"
        
        sheet.merge_cells("I1:J1")
        sheet["I1"] = "6–12 mois"

        sheet.merge_cells("K1:L1")
        sheet["K1"] = "12–18 mois"
        
        sheet.merge_cells("M1:N1")
        sheet["M1"] = "18–24 mois"

        sheet.merge_cells("O1:P1")
        sheet["O1"] = "24-36 mois"

        sheet.merge_cells("Q1:R1")
        sheet["Q1"] = "> 36 mois"

        sheet["C2"] = "QTE"
        sheet["E2"] = "QTE"
        sheet["G2"] = "QTE"
        sheet["K2"] = "QTE"
        sheet["M2"] = "QTE"
        sheet["O2"] = "QTE"
        sheet["Q2"] = "QTE"
        
        sheet["D2"] = "VALEUR ACHAT"
        sheet["F2"] = "VALEUR ACHAT"
        sheet["H2"] = "VALEUR ACHAT"
        sheet["J2"] = "VALEUR ACHAT"
        sheet["L2"] = "VALEUR ACHAT"
        sheet["N2"] = "VALEUR ACHAT"
        sheet["P2"] = "VALEUR ACHAT"
        sheet["R2"] = "VALEUR ACHAT"
        

        

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=medicaments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        workbook.save(response)
        return response
    
    def export_to_excel(self, request, queryset):
        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Medicaments"

        # Header
        headers = ["med_id",
                    "name",
                    "qte_total",
                    "val_total_achat",

                    # < 3 mois
                    "qte_avant_3m",
                    "val_achat_avant_3m",

                    # 3–6 mois
                    "qte_apres_3m_avant_6m",
                    "val_achat_apres_3m_avant_6m",

                    # 6–12 mois
                    "qte_apres_6m_avant_12m",
                    "val_achat_apres_6m_avant_12m",

                    # 12–18 mois
                    "qte_apres_12m_avant_18m",
                    "val_achat_apres_12m_avant_18m",

                    # 18–24 mois
                    "qte_apres_18m_avant_24m",
                    "val_achat_apres_18m_avant_24m",

                    # 24–36 mois
                    "qte_apres_24m_avant_36m",
                    "val_achat_apres_24m_avant_36m",

                    # > 36 mois
                    "qte_apres_36m",
                    "val_achat_apres_36m",]
        sheet.append(headers)

        # Data rows
        for obj in queryset:
            row = [
                    obj.med_id,
                    obj.name,

                    # Totaux
                    obj.qte_total,
                    obj.val_total_achat,

                    # < 3 mois
                    obj.qte_avant_3m,
                    obj.val_achat_avant_3m,

                    # 3–6 mois
                    obj.qte_apres_3m_avant_6m,
                    obj.val_achat_apres_3m_avant_6m,

                    # 6–12 mois
                    obj.qte_apres_6m_avant_12m,
                    obj.val_achat_apres_6m_avant_12m,

                    # 12–18 mois
                    obj.qte_apres_12m_avant_18m,
                    obj.val_achat_apres_12m_avant_18m,

                    # 18–24 mois
                    obj.qte_apres_18m_avant_24m,
                    obj.val_achat_apres_18m_avant_24m,

                    # 24–36 mois
                    obj.qte_apres_24m_avant_36m,
                    obj.val_achat_apres_24m_avant_36m,

                    # > 36 mois
                    obj.qte_apres_36m,
                    obj.val_achat_apres_36m,
                ]

            sheet.append(row)

        sheet.insert_rows(1)
        sheet.insert_rows(1)
        sheet.merge_cells("A1:A2")
        sheet["A1"] = "Med ID"
        sheet.merge_cells("B1:B2")
        sheet["B1"] = "Medicament"
        sheet.merge_cells("C1:D1")
        sheet["C1"] = "TOTAL"
        sheet.merge_cells("E1:F1")
        sheet["E1"] = "< 3 mois"
        
        sheet.merge_cells("G1:H1")
        sheet["G1"] = "3–6 mois"
        
        sheet.merge_cells("I1:J1")
        sheet["I1"] = "6–12 mois"

        sheet.merge_cells("K1:L1")
        sheet["K1"] = "12–18 mois"
        
        sheet.merge_cells("M1:N1")
        sheet["M1"] = "18–24 mois"

        sheet.merge_cells("O1:P1")
        sheet["O1"] = "24-36 mois"

        sheet.merge_cells("Q1:R1")
        sheet["Q1"] = "> 36 mois"

        sheet["C2"] = "QTE"
        sheet["E2"] = "QTE"
        sheet["G2"] = "QTE"
        sheet["K2"] = "QTE"
        sheet["M2"] = "QTE"
        sheet["O2"] = "QTE"
        sheet["Q2"] = "QTE"
        
        sheet["D2"] = "VALEUR ACHAT"
        sheet["F2"] = "VALEUR ACHAT"
        sheet["H2"] = "VALEUR ACHAT"
        sheet["J2"] = "VALEUR ACHAT"
        sheet["L2"] = "VALEUR ACHAT"
        sheet["N2"] = "VALEUR ACHAT"
        sheet["P2"] = "VALEUR ACHAT"
        sheet["R2"] = "VALEUR ACHAT"

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=medicaments_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        workbook.save(response)
        return response

    export_to_excel.short_description = "Exporter la sélection vers Excel"